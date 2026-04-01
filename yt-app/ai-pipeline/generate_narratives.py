"""
Narrative Synthesis Generator — Gemini
═══════════════════════════════════════
Reads videos_with_generated_claims.csv (~16,500 claims from 1500 videos),
sends batches of claims to Gemini to:
  1) First pass: Identify broad narrative themes across ALL claims
  2) Second pass: Assign each claim to a theme
  3) Third pass: For every 50-100 claims in a theme, generate one
     synthesized narrative paragraph

Output: narrative_synthesis.xlsx (formatted like the reference file)

Usage:
    pip install google-generativeai openpyxl pandas
    python generate_narratives.py
"""

import json
import os
import re
import sys
import time
import math
import random
from collections import defaultdict

import pandas as pd
import google.generativeai as genai
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ═══════════════════════════════════════════════════════════════════════════
# CONFIG
# ═══════════════════════════════════════════════════════════════════════════
GOOGLE_API_KEY = #Gemini key here
INPUT_CSV      = "videos_with_generated_claims.csv"
OUTPUT_XLSX    = "narrative_synthesis_output.xlsx"
CHECKPOINT     = "narrative_checkpoint.json"

MODEL_NAME     = "gemini-2.0-flash"     # fast + cheap for bulk
BATCH_SIZE     = 80                     # claims per narrative (50-100 range)
DELAY          = 3                      # seconds between API calls
MAX_RETRIES    = 3
NUM_THEMES     = 12                     # target number of broad themes

genai.configure(api_key=GOOGLE_API_KEY)
model = genai.GenerativeModel(MODEL_NAME)


# ═══════════════════════════════════════════════════════════════════════════
# STEP 1 — Flatten all claims from the wide-format CSV
# ═══════════════════════════════════════════════════════════════════════════

def load_claims(csv_path: str) -> pd.DataFrame:
    """
    Reads the wide CSV and melts it into a long table:
       video_title | channel | claim | narrative_type
    """
    df = pd.read_csv(csv_path)
    claim_cols = sorted([c for c in df.columns if c.startswith("Claim ")],
                        key=lambda x: int(x.split()[-1]))
    narr_cols  = sorted([c for c in df.columns if c.startswith("Narrative ")],
                        key=lambda x: int(x.split()[-1]))

    rows = []
    for _, row in df.iterrows():
        for cc, nc in zip(claim_cols, narr_cols):
            claim = row.get(cc)
            narr  = row.get(nc)
            if pd.notna(claim) and str(claim).strip():
                rows.append({
                    "video_title":    str(row.get("Video Title", "")),
                    "channel":        str(row.get("Channel", "")),
                    "claim":          str(claim).strip(),
                    "narrative_type": str(narr).strip() if pd.notna(narr) else "",
                })
    return pd.DataFrame(rows)


# ═══════════════════════════════════════════════════════════════════════════
# STEP 2 — Call Gemini helper
# ═══════════════════════════════════════════════════════════════════════════

def call_gemini(prompt: str, retries: int = MAX_RETRIES) -> str:
    """Call Gemini with retries and return raw text."""
    for attempt in range(retries):
        try:
            resp = model.generate_content(prompt)
            return resp.text.strip()
        except Exception as e:
            print(f"    ⚠ Gemini error (attempt {attempt+1}/{retries}): {e}")
            if attempt < retries - 1:
                wait = DELAY * (attempt + 2)
                print(f"    Waiting {wait}s...")
                time.sleep(wait)
    return ""


def parse_json_response(text: str) -> dict | list:
    """Robust JSON extraction from Gemini output."""
    # Strip markdown fences
    cleaned = re.sub(r"^```(?:json)?\s*", "", text)
    cleaned = re.sub(r"\s*```$", "", cleaned)
    cleaned = cleaned.strip()
    try:
        return json.loads(cleaned)
    except json.JSONDecodeError:
        # Try to find JSON in the response
        match = re.search(r'[\[{].*[\]}]', cleaned, re.DOTALL)
        if match:
            try:
                return json.loads(match.group())
            except json.JSONDecodeError:
                pass
    return {}


# ═══════════════════════════════════════════════════════════════════════════
# STEP 3 — Discover broad themes
# ═══════════════════════════════════════════════════════════════════════════

def discover_themes(claims_df: pd.DataFrame, num_themes: int = NUM_THEMES) -> list[dict]:
    """
    Sample claims and ask Gemini to identify the major thematic categories.
    Returns list of {"name": "...", "emoji": "...", "description": "..."}
    """
    print("\n🔍 STEP 1: Discovering broad narrative themes...")

    # Sample ~300 claims spread across channels for theme discovery
    sample_size = min(300, len(claims_df))
    sample = claims_df.sample(n=sample_size, random_state=42)
    claims_list = sample["claim"].tolist()

    # Split into 3 sub-batches so we don't blow the context
    chunk_size = len(claims_list) // 3
    chunks = [claims_list[i:i+chunk_size] for i in range(0, len(claims_list), chunk_size)]

    all_theme_suggestions = []
    for i, chunk in enumerate(chunks):
        numbered = "\n".join(f"{j+1}. {c}" for j, c in enumerate(chunk))
        prompt = f"""You are analyzing health & wellness claims from YouTube videos.
Below are {len(chunk)} sample claims. Identify {num_themes} broad thematic categories
that could organize ALL such claims from a dataset of ~16,000 total.

Think about themes like: neuroscience, exercise science, nutrition, mental health,
sleep, supplements, medical practice, longevity, gut health, misinformation debunking, etc.

Return ONLY a JSON array:
[
  {{"name": "Theme Name", "emoji": "🧠", "description": "One sentence describing what claims fall here"}},
  ...
]

CLAIMS:
{numbered}"""
        text = call_gemini(prompt)
        parsed = parse_json_response(text)
        if isinstance(parsed, list):
            all_theme_suggestions.extend(parsed)
        time.sleep(DELAY)

    # Now consolidate the suggestions into final themes
    suggestions_text = json.dumps(all_theme_suggestions, indent=2)
    consolidation_prompt = f"""Below are theme suggestions gathered from sampling health & wellness YouTube claims.
Consolidate these into exactly {num_themes} final, non-overlapping broad themes.
Each theme should be distinct and cover a clear domain.

Return ONLY a JSON array:
[
  {{"name": "Theme Name", "emoji": "🧠", "description": "One-sentence scope of this theme"}},
  ...
]

SUGGESTIONS:
{suggestions_text}"""

    text = call_gemini(consolidation_prompt)
    themes = parse_json_response(text)
    if not isinstance(themes, list) or len(themes) == 0:
        # Fallback themes
        themes = [
            {"name": "Neuroscience & Brain Optimization", "emoji": "🧠", "description": "How the brain works, neuroplasticity, neurochemicals, and cognitive enhancement."},
            {"name": "Exercise, Fitness & Physical Performance", "emoji": "💪", "description": "Training methodology, exercise science, strength, endurance, hypertrophy, and recovery."},
            {"name": "Nutrition, Diet & Metabolic Health", "emoji": "🥗", "description": "What to eat, dietary patterns, macronutrients, micronutrients, and metabolic function."},
            {"name": "Sleep, Recovery & Stress Management", "emoji": "😴", "description": "Sleep optimization, circadian biology, stress response, and recovery protocols."},
            {"name": "Evidence vs. Myth: Debunking Misinformation", "emoji": "🔬", "description": "Challenging popular health narratives and correcting misconceptions."},
            {"name": "Longevity, Aging & Disease Prevention", "emoji": "🧬", "description": "Extending healthspan, preventing chronic disease, and slowing biological aging."},
            {"name": "Mental Health, Psychology & Behavior Change", "emoji": "🧘", "description": "Psychological wellbeing, mental health, habit formation, and motivation."},
            {"name": "Gut Health, Microbiome & Immunity", "emoji": "🌱", "description": "The gut microbiome, immune function, digestion, and gut-brain axis."},
            {"name": "Supplements, Biomarkers & Personalized Health", "emoji": "💊", "description": "Supplements, lab testing, biomarkers, and individualized health optimization."},
            {"name": "Healthcare Systems & Clinical Practice", "emoji": "🏥", "description": "How medicine is practiced, medical education, doctor-patient relationships."},
            {"name": "Hormones & Endocrine Health", "emoji": "⚗️", "description": "Hormonal balance, endocrine disorders, testosterone, estrogen, thyroid."},
            {"name": "Relationships, Social Health & Communication", "emoji": "💬", "description": "Social connections, attachment, communication, and interpersonal dynamics."},
        ]

    print(f"   ✅ Identified {len(themes)} themes:")
    for t in themes:
        print(f"      {t.get('emoji','')} {t['name']}")

    return themes


# ═══════════════════════════════════════════════════════════════════════════
# STEP 4 — Assign each claim to a theme
# ═══════════════════════════════════════════════════════════════════════════

def assign_claims_to_themes(claims_df: pd.DataFrame, themes: list[dict]) -> pd.DataFrame:
    """
    Batch claims and ask Gemini to assign each to the best theme.
    Adds a 'theme' column to the dataframe.
    """
    print(f"\n📂 STEP 2: Assigning {len(claims_df)} claims to themes...")

    theme_names = [t["name"] for t in themes]
    theme_list_str = "\n".join(f"  {i+1}. {t['name']} — {t.get('description','')}"
                               for i, t in enumerate(themes))

    claims_df = claims_df.copy()
    claims_df["theme"] = ""

    # Process in batches of ~150 claims
    assign_batch_size = 150
    total_batches = math.ceil(len(claims_df) / assign_batch_size)

    for batch_idx in range(total_batches):
        start = batch_idx * assign_batch_size
        end   = min(start + assign_batch_size, len(claims_df))
        batch = claims_df.iloc[start:end]

        numbered = "\n".join(f"{i}. {row['claim']}"
                             for i, row in batch.iterrows())

        prompt = f"""Assign each claim below to exactly ONE of these themes.
Return ONLY a JSON object mapping claim index to theme name.
Example: {{"0": "Neuroscience & Brain Optimization", "5": "Nutrition, Diet & Metabolic Health"}}

THEMES:
{theme_list_str}

CLAIMS:
{numbered}"""

        text = call_gemini(prompt)
        mapping = parse_json_response(text)

        if isinstance(mapping, dict):
            assigned = 0
            for idx_str, theme_name in mapping.items():
                try:
                    idx = int(idx_str)
                    if idx in claims_df.index and theme_name in theme_names:
                        claims_df.at[idx, "theme"] = theme_name
                        assigned += 1
                except (ValueError, KeyError):
                    pass
            print(f"   Batch {batch_idx+1}/{total_batches}: assigned {assigned}/{len(batch)} claims")
        else:
            print(f"   Batch {batch_idx+1}/{total_batches}: ⚠ parse failed, using fallback")

        time.sleep(DELAY)

    # Fallback: any unassigned claims get the closest theme by keyword
    unassigned = claims_df[claims_df["theme"] == ""]
    if len(unassigned) > 0:
        print(f"\n   ⚠ {len(unassigned)} claims unassigned — running cleanup pass...")
        # Small batches for cleanup
        for chunk_start in range(0, len(unassigned), 50):
            chunk = unassigned.iloc[chunk_start:chunk_start+50]
            numbered = "\n".join(f"{i}. {row['claim']}" for i, row in chunk.iterrows())
            prompt = f"""Assign each claim to ONE theme. Return JSON: {{"index": "Theme Name", ...}}

THEMES:
{theme_list_str}

CLAIMS:
{numbered}"""
            text = call_gemini(prompt)
            mapping = parse_json_response(text)
            if isinstance(mapping, dict):
                for idx_str, theme_name in mapping.items():
                    try:
                        idx = int(idx_str)
                        if idx in claims_df.index and theme_name in theme_names:
                            claims_df.at[idx, "theme"] = theme_name
                    except (ValueError, KeyError):
                        pass
            time.sleep(DELAY)

    # Final fallback: anything still blank → "Miscellaneous"
    still_empty = claims_df["theme"] == ""
    if still_empty.sum() > 0:
        claims_df.loc[still_empty, "theme"] = theme_names[0]  # default to first
        print(f"   Defaulted {still_empty.sum()} remaining claims to '{theme_names[0]}'")

    # Print distribution
    print("\n   Theme distribution:")
    for theme, count in claims_df["theme"].value_counts().items():
        print(f"      {count:5d} | {theme}")

    return claims_df


# ═══════════════════════════════════════════════════════════════════════════
# STEP 5 — Synthesize narratives (1 per 50-100 claims)
# ═══════════════════════════════════════════════════════════════════════════

def synthesize_narratives(claims_df: pd.DataFrame, themes: list[dict]) -> list[dict]:
    """
    For each theme, batch claims into groups of BATCH_SIZE and generate
    one synthesized narrative per batch.
    Returns list of section dicts for writing to Excel.
    """
    print(f"\n✍️  STEP 3: Generating synthesized narratives (~1 per {BATCH_SIZE} claims)...")

    theme_lookup = {t["name"]: t for t in themes}
    sections = []

    for theme_name in claims_df["theme"].unique():
        theme_claims = claims_df[claims_df["theme"] == theme_name].copy()
        theme_info = theme_lookup.get(theme_name, {"emoji": "📌", "description": ""})

        num_batches = max(1, math.ceil(len(theme_claims) / BATCH_SIZE))
        batch_size_actual = math.ceil(len(theme_claims) / num_batches)

        section = {
            "name": theme_name,
            "emoji": theme_info.get("emoji", "📌"),
            "description": theme_info.get("description", ""),
            "total_claims": len(theme_claims),
            "channels": theme_claims["channel"].nunique(),
            "claims": [],       # all individual claim rows
            "narratives": [],   # synthesized narrative per batch
        }

        for bi in range(num_batches):
            start = bi * batch_size_actual
            end   = min(start + batch_size_actual, len(theme_claims))
            batch = theme_claims.iloc[start:end]

            # Collect individual claims for the Excel listing
            for _, row in batch.iterrows():
                section["claims"].append({
                    "claim": row["claim"],
                    "video": row["video_title"],
                    "channel": row["channel"],
                    "narrative_type": row["narrative_type"],
                })

            # Build the synthesis prompt
            claims_text = "\n".join(f"- [{r['channel']}] {r['claim']}"
                                    for _, r in batch.iterrows())
            prompt = f"""You are a research analyst synthesizing health & wellness claims from YouTube.

Below are {len(batch)} claims from the theme "{theme_name}".
Theme description: {theme_info.get('description', '')}

Write ONE cohesive narrative paragraph (200-350 words) that synthesizes the key
themes, patterns, and insights. Do NOT list claims individually — weave them
into a flowing analytical narrative.

Return ONLY a JSON object:
{{
    "narrative": "Your synthesized paragraph...",
    "dominant_themes": ["theme1", "theme2", "theme3"],
    "key_tensions": "Any contradictions or nuances (or 'None')"
}}

CLAIMS:
{claims_text}"""

            text = call_gemini(prompt)
            result = parse_json_response(text)

            if isinstance(result, dict) and "narrative" in result:
                section["narratives"].append(result)
                print(f"   ✓ {theme_info.get('emoji','')} {theme_name} — batch {bi+1}/{num_batches} ({len(batch)} claims)")
            else:
                section["narratives"].append({
                    "narrative": text if text else "ERROR: Could not generate narrative.",
                    "dominant_themes": [],
                    "key_tensions": "N/A",
                })
                print(f"   ⚠ {theme_name} — batch {bi+1}/{num_batches} (fallback)")

            time.sleep(DELAY)

        sections.append(section)

    return sections


# ═══════════════════════════════════════════════════════════════════════════
# STEP 6 — Write the output Excel (matches reference format)
# ═══════════════════════════════════════════════════════════════════════════

def write_excel(sections: list[dict], total_claims: int, total_videos: int,
                total_channels: int, filepath: str):
    """Write formatted Excel matching the reference narrative_synthesis.xlsx."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Narrative Synthesis"

    # ── Styles ───────────────────────────────────────────────────────────
    title_font     = Font(name="Calibri", bold=True, size=14, color="1F3864")
    subtitle_font  = Font(name="Calibri", size=10, italic=True, color="666666")
    section_font   = Font(name="Calibri", bold=True, size=13, color="1F3864")
    desc_font      = Font(name="Calibri", size=10, italic=True, color="444444")
    header_font    = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
    header_fill    = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    claim_font     = Font(name="Calibri", size=10)
    wrap_top       = Alignment(wrap_text=True, vertical="top")
    center_top     = Alignment(horizontal="center", vertical="top")
    thin_border    = Border(
        left=Side(style="thin", color="B4C6E7"),
        right=Side(style="thin", color="B4C6E7"),
        top=Side(style="thin", color="B4C6E7"),
        bottom=Side(style="thin", color="B4C6E7"),
    )
    narr_fill      = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    narr_font      = Font(name="Calibri", size=10, italic=True, color="375623")

    # ── Column widths ────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 6
    ws.column_dimensions["C"].width = 70
    ws.column_dimensions["D"].width = 45
    ws.column_dimensions["E"].width = 22
    ws.column_dimensions["F"].width = 30

    # ── Title ────────────────────────────────────────────────────────────
    ws.merge_cells("B1:F1")
    ws["B1"] = f"Broad Narrative Analysis — Health & Wellness YouTube ({total_videos} Videos)"
    ws["B1"].font = title_font

    ws.merge_cells("B2:F2")
    ws["B2"] = (f"  {total_claims} claims from {total_videos} videos across "
                f"{total_channels} channels, grouped into {len(sections)} broad narrative themes")
    ws["B2"].font = subtitle_font

    row = 4  # start writing sections

    for sec in sections:
        # ── Section header ───────────────────────────────────────────────
        ws.merge_cells(f"B{row}:F{row}")
        ws[f"B{row}"] = f"{sec['emoji']} {sec['name']}"
        ws[f"B{row}"].font = section_font
        row += 1

        ws.merge_cells(f"B{row}:F{row}")
        ws[f"B{row}"] = (f"  {sec['description']}  |  "
                         f"{sec['total_claims']} claims across {sec['channels']} channels")
        ws[f"B{row}"].font = desc_font
        row += 1

        # ── Column headers ───────────────────────────────────────────────
        for col, h in enumerate(["#", "Claim", "Source Video", "Channel", "Narrative Type"], 2):
            cell = ws.cell(row=row, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
        row += 1

        # ── Claim rows ──────────────────────────────────────────────────
        claim_idx = 0
        narr_idx = 0
        claims_since_narrative = 0

        for c in sec["claims"]:
            claim_idx += 1
            claims_since_narrative += 1

            ws.cell(row=row, column=2, value=claim_idx).font = claim_font
            ws.cell(row=row, column=2).alignment = center_top
            ws.cell(row=row, column=2).border = thin_border

            ws.cell(row=row, column=3, value=c["claim"]).font = claim_font
            ws.cell(row=row, column=3).alignment = wrap_top
            ws.cell(row=row, column=3).border = thin_border

            ws.cell(row=row, column=4, value=c["video"]).font = claim_font
            ws.cell(row=row, column=4).alignment = wrap_top
            ws.cell(row=row, column=4).border = thin_border

            ws.cell(row=row, column=5, value=c["channel"]).font = claim_font
            ws.cell(row=row, column=5).alignment = wrap_top
            ws.cell(row=row, column=5).border = thin_border

            ws.cell(row=row, column=6, value=c["narrative_type"]).font = claim_font
            ws.cell(row=row, column=6).alignment = wrap_top
            ws.cell(row=row, column=6).border = thin_border

            row += 1

            # Insert narrative after each batch of ~BATCH_SIZE claims
            batch_size_actual = math.ceil(sec["total_claims"] / max(1, len(sec["narratives"])))
            if claims_since_narrative >= batch_size_actual and narr_idx < len(sec["narratives"]):
                narr = sec["narratives"][narr_idx]
                narr_text = narr.get("narrative", "")
                themes_str = ", ".join(narr.get("dominant_themes", []))
                tensions = narr.get("key_tensions", "")

                ws.merge_cells(f"B{row}:F{row}")
                full_narr = f"📝 SYNTHESIZED NARRATIVE (claims {claim_idx - claims_since_narrative + 1}–{claim_idx}):\n\n{narr_text}"
                if themes_str:
                    full_narr += f"\n\n🏷️ Dominant themes: {themes_str}"
                if tensions and tensions not in ("None", "N/A", "None identified"):
                    full_narr += f"\n\n⚡ Key tensions: {tensions}"

                cell = ws[f"B{row}"]
                cell.value = full_narr
                cell.font = narr_font
                cell.fill = narr_fill
                cell.alignment = wrap_top
                ws.row_dimensions[row].height = 120
                row += 1

                narr_idx += 1
                claims_since_narrative = 0

        # If there's a remaining narrative that didn't get inserted
        while narr_idx < len(sec["narratives"]):
            narr = sec["narratives"][narr_idx]
            narr_text = narr.get("narrative", "")
            themes_str = ", ".join(narr.get("dominant_themes", []))

            ws.merge_cells(f"B{row}:F{row}")
            full_narr = f"📝 SYNTHESIZED NARRATIVE (final batch):\n\n{narr_text}"
            if themes_str:
                full_narr += f"\n\n🏷️ Dominant themes: {themes_str}"

            cell = ws[f"B{row}"]
            cell.value = full_narr
            cell.font = narr_font
            cell.fill = narr_fill
            cell.alignment = wrap_top
            ws.row_dimensions[row].height = 120
            row += 1
            narr_idx += 1

        row += 1  # blank row between sections

    wb.save(filepath)
    print(f"\n✅ Saved to: {filepath}")


# ═══════════════════════════════════════════════════════════════════════════
# CHECKPOINT helpers (resume if interrupted)
# ═══════════════════════════════════════════════════════════════════════════

def save_checkpoint(data: dict, path: str = CHECKPOINT):
    with open(path, "w") as f:
        json.dump(data, f, indent=2, default=str)
    print(f"   💾 Checkpoint saved → {path}")


def load_checkpoint(path: str = CHECKPOINT) -> dict | None:
    if os.path.exists(path):
        with open(path) as f:
            return json.load(f)
    return None


# ═══════════════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════════════

def main():
    print("=" * 65)
    print("  🎬 Narrative Synthesis Generator — Gemini")
    print("=" * 65)

    # ── Load claims ──────────────────────────────────────────────────────
    print(f"\n📖 Loading claims from {INPUT_CSV}...")
    claims_df = load_claims(INPUT_CSV)
    total_claims   = len(claims_df)
    total_videos   = claims_df["video_title"].nunique()
    total_channels = claims_df["channel"].nunique()
    print(f"   {total_claims} claims from {total_videos} videos, {total_channels} channels")

    # ── Check for checkpoint ─────────────────────────────────────────────
    cp = load_checkpoint()
    if cp and cp.get("step") == "themes_assigned":
        print("\n♻️  Found checkpoint with themes already assigned. Resuming...")
        themes = cp["themes"]
        claims_df["theme"] = cp["theme_assignments"]
    else:
        # ── Discover themes ──────────────────────────────────────────────
        themes = discover_themes(claims_df, NUM_THEMES)

        # ── Assign claims to themes ──────────────────────────────────────
        claims_df = assign_claims_to_themes(claims_df, themes)

        # Save checkpoint after assignment (the expensive part)
        save_checkpoint({
            "step": "themes_assigned",
            "themes": themes,
            "theme_assignments": claims_df["theme"].tolist(),
        })

    # ── Synthesize narratives ────────────────────────────────────────────
    sections = synthesize_narratives(claims_df, themes)

    # ── Write Excel ──────────────────────────────────────────────────────
    write_excel(sections, total_claims, total_videos, total_channels, OUTPUT_XLSX)

    # ── Summary ──────────────────────────────────────────────────────────
    total_narratives = sum(len(s["narratives"]) for s in sections)
    print(f"\n{'═'*65}")
    print(f"  🎉 COMPLETE!")
    print(f"  {total_claims:,} claims → {len(sections)} themes → {total_narratives} narratives")
    print(f"  Output: {OUTPUT_XLSX}")
    print(f"{'═'*65}")


if __name__ == "__main__":
    main()
