import json
import os
import re
import sys
import time
import random
import subprocess

subprocess.check_call([sys.executable, "-m", "pip", "install", "-q", "google-generativeai", "pandas", "youtube-transcript-api", "openpyxl", "yt-dlp"])

import pandas as pd
import google.generativeai as genai
from youtube_transcript_api import YouTubeTranscriptApi
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# --- API KEY SETUP ---
# Option 1 (recommended): set an environment variable so you never type it again
#   Mac/Linux:  export GOOGLE_API_KEY="your_key_here"  (add to ~/.zshrc or ~/.bashrc to persist)
#   Windows:    set GOOGLE_API_KEY=your_key_here
# Option 2: create a .env file in the same folder containing: GOOGLE_API_KEY=your_key_here
# Option 3: just enter it when prompted at runtime

def get_api_key():
    key = os.environ.get("GOOGLE_API_KEY", "")
    if key:
        print("  Using GOOGLE_API_KEY from environment.")
        return key
    if os.path.exists(".env"):
        with open(".env") as f:
            for line in f:
                if line.startswith("GOOGLE_API_KEY="):
                    key = line.strip().split("=", 1)[1].strip().strip('"').strip("'")
                    if key:
                        print("  Using GOOGLE_API_KEY from .env file.")
                        return key
    print("\nNo GOOGLE_API_KEY found in environment or .env file.")
    print("Get your free key at: https://aistudio.google.com/app/apikey")
    key = input("Enter your Google Gemini API key: ").strip()
    if not key:
        print("No API key provided. Exiting.")
        sys.exit(1)
    return key

GOOGLE_API_KEY = get_api_key()
genai.configure(api_key=GOOGLE_API_KEY)

# ==========================================
# CONFIG
# ==========================================
CSV_PATH = "youtube_final_accurate.csv"
URL_COLUMN = "URL"
OUTPUT_PATH = "content_claims.xlsx"
CHECKPOINT_PATH = "checkpoint.json"

# Delay between each transcript fetch (seconds). Increase if still getting banned.
MIN_DELAY = 3
MAX_DELAY = 7

# --- IP BAN FIX ---
# Export your browser cookies using this Chrome extension:
# https://chromewebstore.google.com/detail/get-cookiestxt-locally/cclelndahbckbenkjhflpdbgdldlbecc
# Save the file as cookies.txt in the same folder as this script, then set:
# COOKIES_FILE = "cookies.txt"
COOKIES_FILE = None

# Generic topic prompt — works for any YouTube content
ANALYSIS_TOPIC = """any substantive claims, recommendations, or insights made in the video"""
NARRATIVE_EXAMPLES = """
Examples of narratives: "Evidence-Based", "Contrarian View", "Beginner-Focused",
"Expert Opinion", "Practical Advice", "Theoretical/Academic", "Industry Standard",
"Cautionary", "Motivational", "Data-Driven", etc.
"""
# ==========================================


def extract_video_id(url):
    patterns = [
        r'(?:youtube\.com\/watch\?v=|youtu\.be\/|youtube\.com\/embed\/)([^&\n?#]+)',
        r'youtube\.com\/shorts\/([^&\n?#]+)'
    ]
    for pattern in patterns:
        match = re.search(pattern, url)
        if match:
            return match.group(1)
    return url


def get_video_metadata(video_url):
    try:
        import yt_dlp
        ydl_opts = {'quiet': True, 'no_warnings': True, 'skip_download': True}
        with yt_dlp.YoutubeDL(ydl_opts) as ydl:
            info = ydl.extract_info(video_url, download=False)
            return {
                'title': info.get('title', 'N/A'),
                'channel': info.get('channel', info.get('uploader', 'N/A')),
                'upload_date': info.get('upload_date', 'N/A'),
                'view_count': info.get('view_count', 'N/A'),
            }
    except Exception as e:
        print(f"  Warning: Could not fetch metadata: {e}")
        return {'title': 'N/A', 'channel': 'N/A', 'upload_date': 'N/A', 'view_count': 'N/A'}


def format_date(date_str):
    if date_str and date_str != 'N/A' and len(date_str) == 8:
        return f"{date_str[:4]}-{date_str[4:6]}-{date_str[6:8]}"
    return date_str


def format_views(views):
    if isinstance(views, int):
        if views >= 1_000_000:
            return f"{views/1_000_000:.1f}M"
        elif views >= 1_000:
            return f"{views/1_000:.1f}K"
        return str(views)
    return str(views)


def fetch_youtube_transcript(video_url, retries=3):
    """
    Fetch transcript with retry/backoff logic for IP bans.

    Best fixes for IP bans (in order of effectiveness):
      1. Set COOKIES_FILE to a browser cookies.txt export (see CONFIG above)
      2. Run locally on your home machine instead of a cloud server
      3. Set HTTPS_PROXY env variable to a residential proxy before running
    """
    video_id = extract_video_id(video_url)

    for attempt in range(1, retries + 1):
        try:
            if COOKIES_FILE and os.path.exists(COOKIES_FILE):
                ytt_api = YouTubeTranscriptApi(cookie_path=COOKIES_FILE)
            else:
                ytt_api = YouTubeTranscriptApi()

            transcript = ytt_api.fetch(video_id)
            full_transcript = " ".join([entry.text for entry in transcript])

            sleep_time = random.uniform(MIN_DELAY, MAX_DELAY)
            print(f"  Sleeping {sleep_time:.1f}s...")
            time.sleep(sleep_time)

            return full_transcript, video_id

        except Exception as e:
            err_str = str(e)
            is_ip_ban = any(k in err_str for k in [
                "IPBlocked", "RequestBlocked", "blocking requests", "cloud provider"
            ])

            if is_ip_ban:
                if attempt < retries:
                    wait = 45 * attempt  # 45s, 90s, 135s
                    print(f"  IP ban detected (attempt {attempt}/{retries}). Waiting {wait}s...")
                    time.sleep(wait)
                else:
                    print(f"  IP ban — all {retries} retries exhausted. Skipping.")
                    print(f"  FIX: Export browser cookies to 'cookies.txt' and set COOKIES_FILE above.")
                    return None, None
            else:
                print(f"  Error fetching transcript: {e}")
                return None, None

    return None, None


def analyze_content(text, csv_row_meta=None):
    model = genai.GenerativeModel('gemini-2.5-flash')

    if len(text) > 8000:
        chunk = 2500
        mid = len(text) // 2 - chunk // 2
        text_to_analyze = text[:chunk] + "\n...\n" + text[mid:mid + chunk] + "\n...\n" + text[-chunk:]
    else:
        text_to_analyze = text

    context_hint = ""
    if csv_row_meta:
        context_hint = f"\nVideo Title: {csv_row_meta.get('title', '')}\nChannel: {csv_row_meta.get('channel', '')}\n"

    prompt = f"""
You are an expert content analyst. Your job is to extract meaningful, specific claims and insights from video transcripts.
{context_hint}
Task:
1. Extract AS MANY specific, substantive claims as possible regarding {ANALYSIS_TOPIC}.
   Focus on claims that are actionable, opinionated, or evidence-referenced — not vague filler statements.
2. For EACH claim, identify its "Narrative Leaning" — the stance, framing, or perspective the claim represents.
   {NARRATIVE_EXAMPLES}
3. Extract a minimum of 10 claims; aim for 15-20 if the content supports it. Be thorough and precise.
4. Claims should be self-contained and understandable without needing to watch the video.

Return ONLY valid JSON with no markdown formatting:
{{
    "claims": [
        {{"claim": "Specific claim text here", "narrative": "Narrative leaning label here"}},
        ...
    ]
}}

TRANSCRIPT:
{text_to_analyze}
"""

    try:
        response = model.generate_content(prompt)
        clean_json = response.text.strip()
        if clean_json.startswith("```json"):
            clean_json = clean_json.replace("```json", "").replace("```", "").strip()
        elif clean_json.startswith("```"):
            clean_json = clean_json.replace("```", "").strip()

        parsed = json.loads(clean_json)
        if not parsed.get('claims'):
            print(f"  WARNING: Gemini returned empty claims. Preview: {response.text[:200]}")
        return parsed
    except json.JSONDecodeError as e:
        print(f"  JSON PARSE ERROR: {e}\n  Raw preview: {response.text[:300]}")
        return {"claims": [{"claim": f"JSON parse error: {e}", "narrative": "Error"}]}
    except Exception as e:
        print(f"  GEMINI ERROR: {e}")
        return {"claims": [{"claim": f"Error: {e}", "narrative": "Error"}]}


def load_urls_from_csv(csv_path, url_column):
    df = pd.read_csv(csv_path)
    df.columns = df.columns.str.strip()

    if url_column not in df.columns:
        url_col_candidates = [c for c in df.columns if 'url' in c.lower() or 'link' in c.lower()]
        if url_col_candidates:
            url_column = url_col_candidates[0]
            print(f"  Auto-detected URL column: '{url_column}'")
        else:
            raise ValueError(f"Column '{url_column}' not found. Available: {list(df.columns)}")

    records = []
    for _, row in df.iterrows():
        url = str(row[url_column]).strip()
        if url and url.lower() != 'nan':
            meta = {col: str(row[col]).strip() for col in df.columns if col != url_column}
            meta['url'] = url
            records.append(meta)

    print(f"  Loaded {len(records)} URLs from '{csv_path}'")
    return records


def save_checkpoint(results):
    with open(CHECKPOINT_PATH, "w") as f:
        json.dump(results, f)
    print(f"  Checkpoint saved ({len(results)} videos so far)")


def load_checkpoint():
    if os.path.exists(CHECKPOINT_PATH):
        with open(CHECKPOINT_PATH) as f:
            data = json.load(f)
        print(f"  Resuming from checkpoint: {len(data)} videos already done, skipping those.")
        return data
    return []


def process_videos(records):
    all_results = load_checkpoint()
    processed_urls = {r['url'] for r in all_results}
    skipped_count = 0

    for i, record in enumerate(records, 1):
        url = record['url']

        if url in processed_urls:
            print(f"[{i}/{len(records)}] Already processed, skipping.")
            continue

        print(f"\n{'='*60}")
        print(f"Processing {i}/{len(records)}: {url}")
        print(f"{'='*60}")

        csv_title = record.get('Title') or record.get('title') or ''
        csv_channel = record.get('Channel') or record.get('channel') or ''
        csv_views = record.get('Views') or record.get('views') or ''
        csv_date = record.get('Published') or record.get('Upload Date') or ''

        if csv_title and csv_title != 'N/A':
            print("  Using metadata from CSV...")
            metadata = {'title': csv_title, 'channel': csv_channel,
                        'upload_date': csv_date, 'view_count': csv_views}
            use_raw_meta = True
        else:
            print("  Fetching metadata via yt-dlp...")
            metadata = get_video_metadata(url)
            use_raw_meta = False

        print("  Fetching transcript...")
        transcript, video_id = fetch_youtube_transcript(url)

        if not transcript:
            print("  SKIPPED — No transcript available")
            skipped_count += 1
            continue

        print(f"  Transcript length: {len(transcript):,} chars")
        print("  Analyzing with Gemini...")
        data = analyze_content(transcript, csv_row_meta=metadata)
        claims = data.get('claims', [])
        print(f"  Extracted {len(claims)} claims")

        all_results.append({
            'url': url,
            'title': metadata['title'],
            'channel': metadata['channel'],
            'upload_date': metadata['upload_date'] if use_raw_meta else format_date(metadata['upload_date']),
            'views': str(metadata['view_count']) if use_raw_meta else format_views(metadata['view_count']),
            'claims': claims
        })

        # Auto-save checkpoint every 10 successful videos
        if len(all_results) % 10 == 0:
            save_checkpoint(all_results)

    print(f"\nDone. Processed: {len(all_results)} | Skipped (no transcript): {skipped_count}")
    return all_results


def write_to_excel(results, output_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Content Claims"

    max_claims = max((len(r['claims']) for r in results), default=0)

    header_font = Font(bold=True, color="FFFFFF", size=11, name="Arial")
    header_fill = PatternFill("solid", fgColor="2F5496")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    claim_header_fill = PatternFill("solid", fgColor="4472C4")
    narrative_header_fill = PatternFill("solid", fgColor="5B9BD5")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    base_headers = ["Video Title", "Channel", "Date Published", "Views", "URL"]
    headers = base_headers[:]
    for i in range(1, max_claims + 1):
        headers.append(f"Claim {i}")
        headers.append(f"Narrative {i}")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.alignment = header_align
        cell.border = thin_border
        if col <= len(base_headers):
            cell.fill = header_fill
        elif "Claim" in header:
            cell.fill = claim_header_fill
        else:
            cell.fill = narrative_header_fill

    data_font = Font(name="Arial", size=10)
    wrap_align = Alignment(wrap_text=True, vertical="top")
    claim_fill = PatternFill("solid", fgColor="D6E4F0")
    narrative_fill = PatternFill("solid", fgColor="E2EFDA")

    for row_idx, result in enumerate(results, 2):
        for col, val in enumerate([result['title'], result['channel'], result['upload_date'],
                                    result['views'], result['url']], 1):
            c = ws.cell(row=row_idx, column=col, value=val)
            c.font = data_font
            c.border = thin_border

        for j, claim_data in enumerate(result['claims']):
            claim_col = len(base_headers) + 1 + (j * 2)
            narrative_col = claim_col + 1

            c1 = ws.cell(row=row_idx, column=claim_col, value=claim_data.get('claim', ''))
            c1.font = data_font
            c1.alignment = wrap_align
            c1.fill = claim_fill
            c1.border = thin_border

            c2 = ws.cell(row=row_idx, column=narrative_col, value=claim_data.get('narrative', ''))
            c2.font = data_font
            c2.alignment = wrap_align
            c2.fill = narrative_fill
            c2.border = thin_border

        for col in range(len(result['claims']) * 2 + len(base_headers) + 1, len(headers) + 1):
            ws.cell(row=row_idx, column=col).border = thin_border

    col_widths = {'A': 40, 'B': 20, 'C': 16, 'D': 12, 'E': 40}
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    for i in range(max_claims):
        claim_col_num = len(base_headers) + 1 + (i * 2)
        ws.column_dimensions[get_column_letter(claim_col_num)].width = 55
        ws.column_dimensions[get_column_letter(claim_col_num + 1)].width = 28

    ws.freeze_panes = "A2"
    wb.save(output_path)
    print(f"\nExcel file saved: {output_path}")


# ==========================================
# MAIN
# ==========================================
print("=" * 60)
print("  CONTENT CLAIM EXTRACTION TOOL")
print("=" * 60)
print(f"\nReading URLs from: {CSV_PATH}")

records = load_urls_from_csv(CSV_PATH, URL_COLUMN)
print(f"\n{len(records)} video(s) to process.")

results = process_videos(records)

if results:
    write_to_excel(results, OUTPUT_PATH)
    save_checkpoint(results)
    print("\n" + "=" * 60)
    print("SUMMARY")
    print("=" * 60)
    for r in results:
        print(f"\n  {r['title']}")
        print(f"  Channel: {r['channel']} | Date: {r['upload_date']} | Views: {r['views']}")
        print(f"  Claims: {len(r['claims'])}")
else:
    print("No results to export.")
